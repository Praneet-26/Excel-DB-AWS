import openpyxl as xl
import psycopg2
import configparser


config = configparser.ConfigParser()
config.read('config.ini')
host = config.get('database','host')
database = config.get('database', 'database')
port = config.get('database', 'port')
user = config.get('database', 'user')
password = config.get('database', 'password')

connection = psycopg2.connect(
    database = database,
    host = host,
    port = port,
    user = user,
    password = password
)

cursor = connection.cursor()

workbk_obj = xl.load_workbook("C:\\Users\\Praneet\\OneDrive\\Documents\\SQL Projects\\Excel-Postgres-AWS\\Excel-DB-AWS\\FoodData.xlsx")
sheet_obj = workbk_obj.active

Datatable_Name = workbk_obj.sheetnames

Column_Values = [column.value for column in sheet_obj[1]]

print(Column_Values)

insert_rows = []

for i in sheet_obj.iter_rows(min_row=2, values_only=True):
    insert_rows.append(i)






